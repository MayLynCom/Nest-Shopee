import io
import streamlit as st
import pandas as pd
from processor import processar_tudo

# ── Configuração da página ──────────────────────────────────────────────────
st.set_page_config(
    page_title="Shopee Analytics",
    page_icon="🛒",
    layout="wide",
)


def brl(valor: float) -> str:
    """Formata float para o padrão monetário brasileiro: R$ 1.234,56"""
    formatted = f"{valor:,.2f}"
    return "R$ " + formatted.replace(",", "X").replace(".", ",").replace("X", ".")


def _gerar_excel(
    df_filtrado: pd.DataFrame,
    gmv: float,
    receita_ads: float,
    investimento_ads: float,
    tacos: float,
) -> bytes:
    """Gera um arquivo Excel com aba de métricas e aba de produtos."""
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        # Aba 1 — Resumo de métricas
        resumo = pd.DataFrame({
            "Métrica": ["GMV (Faturamento Total)", "Receita ADS", "Investimento ADS", "TACOS (%)"],
            "Valor": [gmv, receita_ads, investimento_ads, round(tacos, 2)],
        })
        resumo.to_excel(writer, sheet_name="Resumo", index=False)

        # Aba 2 — Produtos (seleção atual do filtro de curvas)
        df_export = pd.DataFrame({
            "ID do Produto": df_filtrado["ID do Produto"].values,
            "Nome do Produto": df_filtrado["Nome"].values,
            "Curva": df_filtrado["Curva"].values,
            "Unidades Vendidas": df_filtrado["Unidades Vendidas"].values,
            "Ticket Médio (R$)": df_filtrado["Ticket Médio"].round(2).values,
            "Faturamento Total (R$)": df_filtrado["Faturamento"].round(2).values,
            "ADS": df_filtrado["ADS"].values,
        })
        df_export.to_excel(writer, sheet_name="Produtos", index=False)

        # Aplicar cor laranja nas linhas de produtos em ADS
        wb = writer.book
        ws = wb["Produtos"]
        from openpyxl.styles import PatternFill, Font
        fill_ads = PatternFill(start_color="FFE0D6", end_color="FFE0D6", fill_type="solid")
        font_ads = Font(color="BF3A1E", bold=True)
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            if row[-1].value == "Sim":
                for cell in row:
                    cell.fill = fill_ads
                    cell.font = font_ads

        # Ajustar largura das colunas automaticamente
        for sheet in wb.sheetnames:
            ws_cur = wb[sheet]
            for col in ws_cur.columns:
                max_len = max((len(str(c.value)) if c.value else 0) for c in col)
                ws_cur.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)

    return output.getvalue()


# ── CSS com identidade visual Shopee ────────────────────────────────────────
st.markdown(
    """
    <style>
    :root {
        --shopee-orange: #EE4D2D;
        --shopee-orange-light: #FF7337;
        --shopee-bg: #F5F5F5;
        --shopee-card-bg: #FFFFFF;
        --shopee-text: #333333;
        --shopee-border: #E8E8E8;
    }

    .stApp { background-color: var(--shopee-bg); }

    /* Header */
    .shopee-header {
        background: linear-gradient(135deg, #EE4D2D 0%, #FF7337 100%);
        padding: 18px 32px;
        border-radius: 12px;
        margin-bottom: 24px;
        display: flex;
        align-items: center;
        gap: 14px;
        box-shadow: 0 4px 16px rgba(238,77,45,0.25);
    }
    .shopee-header h1 { color: white; margin: 0; font-size: 28px; font-weight: 800; letter-spacing: -0.5px; }
    .shopee-header p  { color: rgba(255,255,255,0.85); margin: 2px 0 0 0; font-size: 14px; }
    .shopee-logo      { font-size: 38px; line-height: 1; }

    /* Metric cards */
    .metric-card {
        background: var(--shopee-card-bg);
        border-radius: 10px;
        padding: 20px 22px;
        border: 1px solid var(--shopee-border);
        box-shadow: 0 2px 8px rgba(0,0,0,0.06);
        text-align: center;
        height: 100%;
    }
    .metric-label {
        font-size: 11px;
        color: #888;
        font-weight: 700;
        text-transform: uppercase;
        letter-spacing: 0.8px;
        margin-bottom: 8px;
    }
    .metric-value      { font-size: 24px; font-weight: 800; color: var(--shopee-orange); }
    .metric-value-dark { font-size: 24px; font-weight: 800; color: var(--shopee-text); }

    /* Upload title */
    .upload-title {
        font-size: 15px;
        font-weight: 700;
        color: var(--shopee-text);
        margin: 0 0 10px 0;
        padding-bottom: 8px;
        border-bottom: 2px solid var(--shopee-orange);
        display: inline-block;
    }

    /* File uploader — forçar texto preto para fundo branco */
    [data-testid="stFileUploader"] {
        background-color: #FFFFFF !important;
        border-radius: 8px !important;
    }
    [data-testid="stFileUploader"] section {
        background-color: #F9F9F9 !important;
        border: 1.5px dashed #D0D0D0 !important;
        border-radius: 8px !important;
    }
    [data-testid="stFileUploader"] section * {
        color: #333333 !important;
    }
    [data-testid="stFileUploader"] section svg {
        fill: #888888 !important;
    }
    /* Nomes dos arquivos enviados */
    [data-testid="stFileUploader"] [data-testid="stFileUploaderFile"] {
        background-color: #FFF5F3 !important;
        border: 1px solid #FECACA !important;
        border-radius: 6px !important;
    }
    [data-testid="stFileUploader"] [data-testid="stFileUploaderFile"] * {
        color: #333333 !important;
    }
    [data-testid="stFileUploaderFileName"] {
        color: #333333 !important;
        font-weight: 600 !important;
    }

    /* Botão processar */
    .stButton > button {
        background: linear-gradient(135deg, #EE4D2D, #FF7337) !important;
        color: white !important;
        border: none !important;
        border-radius: 8px !important;
        font-size: 16px !important;
        font-weight: 700 !important;
        padding: 12px 36px !important;
        width: 100% !important;
        transition: all 0.2s ease !important;
        box-shadow: 0 4px 12px rgba(238,77,45,0.35) !important;
    }
    .stButton > button:hover {
        transform: translateY(-1px) !important;
        box-shadow: 0 6px 18px rgba(238,77,45,0.45) !important;
    }

    /* Section titles */
    .section-title {
        font-size: 17px;
        font-weight: 700;
        color: var(--shopee-text);
        margin: 24px 0 12px 0;
        padding-left: 10px;
        border-left: 4px solid var(--shopee-orange);
    }

    /* Badge contador de arquivos */
    .file-badge {
        display: inline-block;
        background: #EE4D2D;
        color: white;
        font-size: 12px;
        font-weight: 700;
        padding: 2px 10px;
        border-radius: 20px;
        margin-top: 6px;
    }

    .stDataFrame { border-radius: 10px; overflow: hidden; }

    /* Tag do multiselect laranja */
    .stMultiSelect [data-baseweb="tag"] {
        background-color: var(--shopee-orange) !important;
    }
    </style>
    """,
    unsafe_allow_html=True,
)

# ── Header ──────────────────────────────────────────────────────────────────
st.markdown(
    """
    <div class="shopee-header">
        <div class="shopee-logo">🛒</div>
        <div>
            <h1>Shopee Analytics</h1>
            <p>Análise de desempenho de produtos e campanhas ADS</p>
        </div>
    </div>
    """,
    unsafe_allow_html=True,
)

# ── Upload de arquivos ───────────────────────────────────────────────────────
st.markdown('<div class="section-title">Importar Arquivos</div>', unsafe_allow_html=True)

col1, col2, col3 = st.columns(3)

with col1:
    st.markdown('<p class="upload-title">📊 Relatório de Produtos</p>', unsafe_allow_html=True)
    file_produtos = st.file_uploader(
        "xlsx de métricas de produtos",
        type=["xlsx"],
        key="file_produtos",
        label_visibility="collapsed",
    )
    if file_produtos:
        st.markdown('<span class="file-badge">✔ 1 arquivo carregado</span>', unsafe_allow_html=True)
    else:
        st.caption("Arquivo xlsx — aba 'Produtos com Melhor Desempenho'")

with col2:
    st.markdown('<p class="upload-title">📣 Relatório Geral de ADS</p>', unsafe_allow_html=True)
    file_ads = st.file_uploader(
        "CSV Dados Gerais de Anúncios Shopee",
        type=["csv"],
        key="file_ads",
        label_visibility="collapsed",
    )
    if file_ads:
        st.markdown('<span class="file-badge">✔ 1 arquivo carregado</span>', unsafe_allow_html=True)
    else:
        st.caption("CSV — Dados Gerais de Anúncios Shopee")

with col3:
    st.markdown('<p class="upload-title">📁 Grupos de ADS (opcional)</p>', unsafe_allow_html=True)
    files_grupos = st.file_uploader(
        "CSVs de grupo de anúncios",
        type=["csv"],
        key="files_grupos",
        accept_multiple_files=True,
        label_visibility="collapsed",
    )
    if files_grupos:
        qtd = len(files_grupos)
        st.markdown(
            f'<span class="file-badge">✔ {qtd} arquivo{"s" if qtd > 1 else ""} carregado{"s" if qtd > 1 else ""}</span>',
            unsafe_allow_html=True,
        )
    else:
        st.caption("Envie 1 ou mais CSVs de grupos de anúncios (sem limite)")

# ── Botão processar ──────────────────────────────────────────────────────────
st.markdown("<br>", unsafe_allow_html=True)
_, col_btn, _ = st.columns([2, 2, 2])
with col_btn:
    processar = st.button("⚡ Processar Arquivos", use_container_width=True)

# ── Ao clicar em Processar: rodar e salvar no session_state ─────────────────
if processar:
    if not file_produtos or not file_ads:
        st.warning(
            "⚠️ Envie os dois arquivos obrigatórios: **Relatório de Produtos** e **Relatório de ADS**."
        )
        st.stop()

    with st.spinner("Processando arquivos..."):
        try:
            resultado = processar_tudo(
                file_produtos=file_produtos,
                file_ads=file_ads,
                files_grupos=files_grupos if files_grupos else None,
            )
        except Exception as e:
            st.error(f"Erro ao processar os arquivos: {e}")
            st.stop()

    # Salvar no session_state para persistir ao usar filtros
    st.session_state["resultado"] = resultado

# ── Exibir resultados (lê do session_state, mantido mesmo ao filtrar) ────────
if "resultado" in st.session_state:
    resultado = st.session_state["resultado"]
    df: pd.DataFrame = resultado["df"]
    gmv: float = resultado["gmv"]
    receita_ads: float = resultado["receita_ads"]
    investimento_ads: float = resultado["investimento_ads"]
    tacos: float = resultado["tacos"]

    # ── Métricas principais ───────────────────────────────────────────────
    st.markdown('<div class="section-title">Visão Geral</div>', unsafe_allow_html=True)

    m1, m2, m3, m4 = st.columns(4)
    with m1:
        st.markdown(
            f'<div class="metric-card"><div class="metric-label">GMV (Faturamento Total)</div>'
            f'<div class="metric-value">{brl(gmv)}</div></div>',
            unsafe_allow_html=True,
        )
    with m2:
        st.markdown(
            f'<div class="metric-card"><div class="metric-label">Receita ADS</div>'
            f'<div class="metric-value-dark">{brl(receita_ads)}</div></div>',
            unsafe_allow_html=True,
        )
    with m3:
        st.markdown(
            f'<div class="metric-card"><div class="metric-label">Investimento ADS</div>'
            f'<div class="metric-value-dark">{brl(investimento_ads)}</div></div>',
            unsafe_allow_html=True,
        )
    with m4:
        st.markdown(
            f'<div class="metric-card"><div class="metric-label">TACOS</div>'
            f'<div class="metric-value">{tacos:.2f}%</div></div>',
            unsafe_allow_html=True,
        )

    # ── Curva ABC — resumo + filtro ───────────────────────────────────────
    st.markdown('<div class="section-title">Curva ABC</div>', unsafe_allow_html=True)

    col_curvas, col_filtro = st.columns([3, 1])
    with col_curvas:
        ca1, ca2, ca3 = st.columns(3)
        for col, letra, cor, label in [
            (ca1, "A", "#22c55e", "🟢 Curva A — 80% do GMV"),
            (ca2, "B", "#3b82f6", "🔵 Curva B — 15% do GMV"),
            (ca3, "C", "#f59e0b", "🟡 Curva C — 5% do GMV"),
        ]:
            subset = df[df["Curva"] == letra]
            fat = subset["Faturamento"].sum()
            with col:
                st.markdown(
                    f'<div class="metric-card">'
                    f'<div class="metric-label" style="color:{cor};">{label}</div>'
                    f'<div class="metric-value" style="color:{cor};">{len(subset)} produtos</div>'
                    f'<div style="font-size:13px;color:#666;margin-top:4px;">{brl(fat)}</div>'
                    f"</div>",
                    unsafe_allow_html=True,
                )

    with col_filtro:
        st.markdown("<br>", unsafe_allow_html=True)
        curvas_selecionadas = st.multiselect(
            "Filtrar por Curva",
            options=["A", "B", "C"],
            default=["A", "B", "C"],
            help="Selecione as curvas que deseja visualizar na tabela",
            key="filtro_curvas",
        )

    # ── Tabela de produtos ────────────────────────────────────────────────
    st.markdown('<div class="section-title">Produtos</div>', unsafe_allow_html=True)

    if not curvas_selecionadas:
        st.info("Selecione ao menos uma curva no filtro acima para exibir os produtos.")
    else:
        df_filtrado = df[df["Curva"].isin(curvas_selecionadas)].copy()

        df_exibir = pd.DataFrame({
            "ID do Produto": df_filtrado["ID do Produto"].values,
            "Nome do Produto": df_filtrado["Nome"].values,
            "Curva": df_filtrado["Curva"].values,
            "Unidades Vendidas": df_filtrado["Unidades Vendidas"].values,
            "Ticket Médio": df_filtrado["Ticket Médio"].round(2).values,
            "Faturamento Total": df_filtrado["Faturamento"].round(2).values,
            "ADS": df_filtrado["ADS"].values,
        })

        def estilizar_linha(row):
            if row["ADS"] == "Sim":
                return ["background-color: #FFF0EB; color: #BF3A1E; font-weight: 600"] * len(row)
            return [""] * len(row)

        styled = (
            df_exibir.style
            .apply(estilizar_linha, axis=1)
            .format({
                "Ticket Médio": lambda x: brl(x) if isinstance(x, (int, float)) else x,
                "Faturamento Total": lambda x: brl(x) if isinstance(x, (int, float)) else x,
            })
        )

        st.dataframe(
            styled,
            use_container_width=True,
            height=530,
            column_config={
                "ID do Produto": st.column_config.TextColumn("ID do Produto", width=140),
                "Nome do Produto": st.column_config.TextColumn("Nome do Produto", width=300),
                "Curva": st.column_config.TextColumn("Curva", width=65),
                "Unidades Vendidas": st.column_config.NumberColumn("Unid. Vendidas", width=110),
                "Ticket Médio": st.column_config.TextColumn("Ticket Médio", width=130),
                "Faturamento Total": st.column_config.TextColumn("Faturamento Total", width=145),
                "ADS": st.column_config.TextColumn("ADS", width=60),
            },
            hide_index=True,
        )

        ads_count = int((df_filtrado["ADS"] == "Sim").sum())
        col_legenda, col_download = st.columns([3, 1])
        with col_legenda:
            st.markdown(
                f'<div style="margin-top:8px; font-size:13px; color:#888;">'
                f'<span style="background:#FFF0EB; color:#EE4D2D; padding:2px 9px; '
                f'border-radius:4px; font-weight:700; border:1px solid #FECACA;">laranja</span>'
                f" = produto em campanha ADS &nbsp;|&nbsp; "
                f"<b>{ads_count}</b> de <b>{len(df_filtrado)}</b> produtos nesta seleção estão em ADS"
                f"</div>",
                unsafe_allow_html=True,
            )
        with col_download:
            excel_bytes = _gerar_excel(df_filtrado, gmv, receita_ads, investimento_ads, tacos)
            st.download_button(
                label="⬇️ Baixar Excel",
                data=excel_bytes,
                file_name="shopee_analytics.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )

# ── Rodapé ───────────────────────────────────────────────────────────────────
st.markdown("<br><br>", unsafe_allow_html=True)
st.markdown(
    '<div style="text-align:center; color:#bbb; font-size:12px; padding:12px 0;">'
    "Shopee Analytics • Análise de desempenho de produtos e ADS"
    "</div>",
    unsafe_allow_html=True,
)
